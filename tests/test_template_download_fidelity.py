from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook

from bling_app_zero.core.template_download_exporter import (
    build_template_download_bytes,
    can_export_from_template,
    mime_for_template_output,
    output_name_for_template,
)


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    ws_info = workbook.active
    ws_info.title = 'INSTRUCOES'
    ws_info['A1'] = 'Nao apagar esta aba'
    ws_info['A2'] = 'Modelo generico de qualquer portal ou sistema'

    ws_data = workbook.create_sheet('DADOS')
    ws_data['A1'] = 'Planilha modelo original'
    ws_data.append([])
    ws_data.append(['Codigo Final', 'Nome Final', 'Loja'])
    ws_data.append(['EXEMPLO-1', 'Produto Exemplo', 'Loja Exemplo'])

    ws_aux = workbook.create_sheet('base_modelo')
    ws_aux['A1'] = 'Valores auxiliares preservados'

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_xlsx_template_download_preserves_workbook_and_fills_mapped_rows() -> None:
    template_bytes = _workbook_bytes()
    df = pd.DataFrame(
        {
            'Codigo Final': ['P001', 'P002'],
            'Nome Final': ['Caneca', 'Mouse'],
            'Loja': ['Mega Center', 'Mega Center'],
        }
    )

    result_bytes = build_template_download_bytes(
        template_bytes=template_bytes,
        template_name='modelo_generico.xlsx',
        df=df,
    )

    workbook = load_workbook(BytesIO(result_bytes), data_only=False)
    assert workbook.sheetnames == ['INSTRUCOES', 'DADOS', 'base_modelo']
    assert workbook['INSTRUCOES']['A1'].value == 'Nao apagar esta aba'
    assert workbook['base_modelo']['A1'].value == 'Valores auxiliares preservados'

    ws_data = workbook['DADOS']
    assert [ws_data.cell(row=3, column=idx).value for idx in range(1, 4)] == ['Codigo Final', 'Nome Final', 'Loja']
    assert [ws_data.cell(row=4, column=idx).value for idx in range(1, 4)] == ['P001', 'Caneca', 'Mega Center']
    assert [ws_data.cell(row=5, column=idx).value for idx in range(1, 4)] == ['P002', 'Mouse', 'Mega Center']
    assert ws_data['A6'].value is None
    workbook.close()


def test_template_download_keeps_original_extension_and_mime() -> None:
    template_bytes = _workbook_bytes()

    assert can_export_from_template('modelo_generico.xlsx', template_bytes) is True
    assert output_name_for_template('modelo_generico.xlsx') == 'modelo_generico_preenchido.xlsx'
    assert mime_for_template_output('modelo_generico.xlsx') == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def test_csv_template_download_keeps_separator_from_original_model() -> None:
    template_bytes = 'Codigo Final;Nome Final;Loja\nEXEMPLO;Produto Exemplo;Loja Exemplo\n'.encode('utf-8-sig')
    df = pd.DataFrame(
        {
            'Codigo Final': ['P001'],
            'Nome Final': ['Caneca'],
            'Loja': ['Mega Center'],
        }
    )

    result_bytes = build_template_download_bytes(
        template_bytes=template_bytes,
        template_name='modelo_generico.csv',
        df=df,
    )

    text = result_bytes.decode('utf-8-sig')
    assert 'Codigo Final;Nome Final;Loja' in text
    assert 'P001;Caneca;Mega Center' in text
    assert ',' not in text.splitlines()[0]
