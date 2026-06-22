from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook

from bling_app_zero.core.files import read_uploaded_file


class UploadedBytes:
    def __init__(self, name: str, data: bytes) -> None:
        self.name = name
        self._data = data

    def getvalue(self) -> bytes:
        return self._data


def _save(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_detects_header_below_instruction_rows() -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = 'Modelo'
    ws['A1'] = 'Planilha de importacao - nao apague as instrucoes'
    ws['A2'] = 'Preencha os campos abaixo conforme a origem'
    ws.append(['Codigo Final', 'Nome Final', 'Preco Final'])
    ws.append(['EXEMPLO-1', 'Produto Exemplo', '999'])

    df = read_uploaded_file(UploadedBytes('modelo_com_texto_acima.xlsx', _save(workbook)))

    assert isinstance(df, pd.DataFrame)
    assert df.columns.tolist() == ['Codigo Final', 'Nome Final', 'Preco Final']
    assert df.iloc[0].tolist() == ['EXEMPLO-1', 'Produto Exemplo', '999']


def test_detects_best_sheet_when_first_sheet_is_instructions() -> None:
    workbook = Workbook()
    ws_help = workbook.active
    ws_help.title = 'INSTRUCOES'
    ws_help['A1'] = 'Leia antes de preencher esta planilha'
    ws_help['A2'] = 'Esta aba contem apenas textos explicativos'

    ws_model = workbook.create_sheet('PRODUTOS')
    ws_model['A1'] = 'Area de dados'
    ws_model.append([])
    ws_model.append(['SKU', 'Produto', 'Estoque', 'Preco'])
    ws_model.append(['A1', 'Caneca', '10', '12.90'])
    ws_model.append(['A2', 'Mouse', '5', '49.90'])

    df = read_uploaded_file(UploadedBytes('modelo_multiplas_abas.xlsx', _save(workbook)))

    assert df.columns.tolist() == ['SKU', 'Produto', 'Estoque', 'Preco']
    assert df.shape == (2, 4)
    assert df['SKU'].tolist() == ['A1', 'A2']


def test_accepts_header_only_sheet_as_model_contract() -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = 'MODELO'
    ws['A4'] = 'Codigo Final'
    ws['B4'] = 'Nome Final'
    ws['C4'] = 'Loja'

    df = read_uploaded_file(UploadedBytes('modelo_header_only_linha_4.xlsx', _save(workbook)))

    assert df.empty
    assert df.columns.tolist() == ['Codigo Final', 'Nome Final', 'Loja']


def test_ignores_instruction_sheet_with_many_texts_and_picks_wide_header() -> None:
    workbook = Workbook()
    ws_help = workbook.active
    ws_help.title = 'Leia-me'
    ws_help['A1'] = 'Este arquivo contem diversas regras de preenchimento para o usuario seguir antes de importar.'
    ws_help['A2'] = 'Nao use esta aba como dados finais.'

    ws_data = workbook.create_sheet('Dados')
    ws_data.append(['Codigo', 'Descricao', 'GTIN/EAN', 'Marca', 'Categoria', 'Preco', 'Estoque'])

    df = read_uploaded_file(UploadedBytes('modelo_generico_sem_dados.xlsx', _save(workbook)))

    assert df.empty
    assert df.columns.tolist() == ['Codigo', 'Descricao', 'GTIN/EAN', 'Marca', 'Categoria', 'Preco', 'Estoque']
