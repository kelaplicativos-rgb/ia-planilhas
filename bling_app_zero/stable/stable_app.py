from __future__ import annotations

from html import escape
from io import BytesIO
from math import ceil
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from bling_app_zero.core.file_reader import read_uploaded_table
from bling_app_zero.stable.session_vault import guardar_df, limpar_vault, restaurar_chaves_df, restaurar_df
from bling_app_zero.stable.supplier_upload_v2 import render_supplier_upload_v2
from bling_app_zero.ui.app_helpers import dataframe_para_csv_bytes
from bling_app_zero.ui.flash_amplo_execution import executar_flash_amplo_pagina_por_pagina
from bling_app_zero.ui.mapeamento.mapping_engine import analyze_mapping
from bling_app_zero.ui.mapeamento.source_columns import normalizar_nome_coluna
from bling_app_zero.ui.mapeamento.value_guard import clean_invalid_preview_mappings

FLASH_MAX_PRODUCTS = 5000
BLING_MAX_IMPORT_ROWS = 1000
CAD_COLS = ["Código", "Descrição", "Descrição complementar", "Unidade", "GTIN/EAN", "Preço unitário", "Marca", "Categoria", "URL imagens externas", "Link Externo", "URL do Produto"]
EST_COLS = ["Código", "Descrição", "GTIN/EAN", "Depósito", "Estoque", "Quantidade"]
GS1_PREFIXOS_INVALIDOS_BLOQUEADOS = {"665", "684", "782", "852"}
MAPPING_UI_VERSION = "2026-05-07-auto-model-v1"


def _read_upload(file) -> pd.DataFrame | None:
    if file is None:
        return None
    return read_uploaded_table(file).dataframe.fillna("")


def _has_df(df) -> bool:
    return isinstance(df, pd.DataFrame) and not df.empty and len(df.columns) > 0


def _has_model_columns(df) -> bool:
    return isinstance(df, pd.DataFrame) and len([c for c in df.columns if str(c).strip()]) > 0


def _read_bling_model_upload(file) -> pd.DataFrame | None:
    if file is None:
        return None
    try:
        df = _read_upload(file)
        if _has_model_columns(df):
            return df
    except Exception:
        pass
    try:
        file.seek(0)
        raw = file.read()
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cols = [str(v).strip() for v in row if str(v or "").strip()]
                if len(cols) >= 2:
                    return pd.DataFrame(columns=cols)
    except Exception:
        return None
    return None


def _norm(v) -> str:
    return normalizar_nome_coluna(v)


def _detect_tipo_by_model(modelo: pd.DataFrame | None) -> str:
    if not _has_model_columns(modelo):
        return "cadastro"
    nomes = {_norm(c) for c in modelo.columns if str(c).strip()}
    texto = " | ".join(sorted(nomes))
    estoque_score = 0
    cadastro_score = 0
    for termo in ("deposito", "estoque", "quantidade", "saldo", "balanco"):
        if termo in texto:
            estoque_score += 2
    for termo in ("descricao complementar", "url imagens externas", "categoria", "marca", "ncm", "peso bruto", "peso liquido", "unidade"):
        if termo in texto:
            cadastro_score += 2
    return "estoque" if estoque_score > cadastro_score else "cadastro"


def _cols(tipo: str, modelo: pd.DataFrame | None) -> list[str]:
    if _has_model_columns(modelo):
        return [str(c).strip() for c in modelo.columns if str(c).strip()]
    return EST_COLS if tipo == "estoque" else CAD_COLS


def _digits(value: object) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _is_ncm_col(col: object) -> bool:
    n = _norm(col)
    return n == "ncm" or "classificacao fiscal" in n or "class fiscal" in n or "codigo ncm" in n


def _clean_ncm_value(value: object) -> str:
    digits = _digits(value)
    return digits if len(digits) == 8 else ""


def _limpar_ncms_invalidos_automatico(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df, 0
    out = df.copy().fillna("")
    total = 0
    for col in out.columns:
        if not _is_ncm_col(col):
            continue
        antes = out[col].astype(str).fillna("")
        depois = antes.map(_clean_ncm_value)
        total += int((antes.str.strip() != depois.astype(str).str.strip()).sum())
        out[col] = depois
    return out.fillna(""), total


def _is_gtin_col(col: object) -> bool:
    n = _norm(col)
    return "gtin" in n or "ean" in n or "codigo de barras" in n or "cod barras" in n


def _gtin_check_digit_ok(digits: str) -> bool:
    if len(digits) not in {8, 12, 13, 14} or not digits.isdigit():
        return False
    total = 0
    for idx, char in enumerate(digits[:-1][::-1]):
        total += int(char) * (3 if idx % 2 == 0 else 1)
    expected = (10 - (total % 10)) % 10
    return int(digits[-1]) == expected


def _clean_gtin_value(value: object) -> str:
    digits = _digits(value)
    if not digits:
        return ""
    if len(digits) not in {8, 12, 13, 14}:
        return ""
    if len(digits) >= 3 and digits[:3] in GS1_PREFIXOS_INVALIDOS_BLOQUEADOS:
        return ""
    if not _gtin_check_digit_ok(digits):
        return ""
    return digits


def _limpar_gtins_invalidos_automatico(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df, 0
    out = df.copy().fillna("")
    total = 0
    for col in out.columns:
        if not _is_gtin_col(col):
            continue
        antes = out[col].astype(str).fillna("")
        depois = antes.map(_clean_gtin_value)
        total += int((antes.str.strip() != depois.astype(str).str.strip()).sum())
        out[col] = depois
    return out.fillna(""), total


def _is_tipo_item_col(col: object) -> bool:
    n = _norm(col)
    return n in {"tipo do item", "tipo item", "tipo de item"}


def _is_unidade_medida_col(col: object) -> bool:
    n = _norm(col)
    return n in {"unidade", "unidade de medida", "unidade medida", "unid", "un"}


def _is_auto_map_blocked_target(col: object) -> bool:
    return _is_tipo_item_col(col) or _is_unidade_medida_col(col)


def _is_deposito_col(col: object) -> bool:
    return "deposito" in _norm(col)


def _is_imagem_col(col: object) -> bool:
    n = _norm(col)
    return "imagem" in n or "image" in n or ("url" in n and ("foto" in n or "externa" in n))


def _is_imagem_target(col: object) -> bool:
    n = _norm(col)
    return "imagem" in n or "image" in n


def _is_descricao_col(col: object) -> bool:
    n = _norm(col)
    return "descricao" in n or n in {"nome", "produto"}


def _is_preco_venda_col(col: object) -> bool:
    n = _norm(col)
    return ("preco" in n or "valor" in n) and "custo" not in n and "compra" not in n


def _dedupe_source_columns(columns) -> list[str]:
    saida: list[str] = []
    vistos: set[str] = set()
    for col in columns:
        nome = str(col).strip()
        if not nome:
            continue
        chave = _norm(nome).replace(" ", "_")
        if chave in vistos:
            continue
        vistos.add(chave)
        saida.append(nome)
    return saida


def _column_has_data(df: pd.DataFrame, col: str) -> bool:
    if not _has_df(df) or col not in df.columns:
        return False
    try:
        return bool(df[col].astype(str).str.strip().ne("").any())
    except Exception:
        return False


def _is_blocked_bling_only_column(col: object) -> bool:
    n = _norm(col)
    blocked_exact = {"gtin ean da embalagem", "frete gratis", "preco unitario obrigatorio", "preco unitario obrigatorio obrigatorio", "categoria do produto"}
    return n in blocked_exact or "obrigatorio" in n or "obrigatoria" in n


def _source_options_for_target(target: str, df: pd.DataFrame, model_cols: list[str]) -> list[str]:
    raw_cols = _dedupe_source_columns(df.columns if _has_df(df) else [])
    cols = [c for c in raw_cols if not _is_blocked_bling_only_column(c) and _column_has_data(df, c)]
    if _is_imagem_target(target):
        cols = [c for c in cols if _is_imagem_col(c)]
    if _is_ncm_col(target):
        cols = [c for c in cols if _is_ncm_col(c)]
    return [""] + cols


def _auto_map_100(target: str, options: list[str], df: pd.DataFrame) -> str:
    if _is_auto_map_blocked_target(target):
        return ""
    target_norm = _norm(target)
    if not target_norm:
        return ""
    candidatos = [opt for opt in options if opt and _norm(opt) == target_norm and _column_has_data(df, opt)]
    validos: list[str] = []
    for opt in candidatos:
        analise = analyze_mapping(df, target, opt)
        if analise.status == "valid" and analise.confidence >= 0.70:
            validos.append(opt)
    return validos[0] if len(validos) == 1 else ""


def _is_image_url_good(url: str) -> bool:
    u = str(url or "").strip().strip('"').strip("'")
    low = u.lower()
    if not u.startswith(("http://", "https://")):
        return False
    if "{" in u or "}" in u or "\\" in u or "original" in low or "thumbnail" in low:
        return False
    if "/produto/" in low and "storage/" not in low and "product_images" not in low:
        return False
    if "rs:fit" in low:
        for tiny in ("120:120", "256:256", "400:400", "50:50"):
            if tiny in low:
                return False
    if any(bad in low for bad in ("logo", "sprite", "placeholder", "sem-imagem", "no-image", "favicon")):
        return False
    if not any(ext in low for ext in (".jpg", ".jpeg", ".png", ".webp")):
        return False
    return True


def _clean_image_urls_value(value: object) -> str:
    raw = str(value or "").replace("\r", "|").replace("\n", "|").replace(";", "|").replace(",http", "|http")
    parts = [p.strip() for p in raw.split("|") if p.strip()]
    good: list[str] = []
    seen: set[str] = set()
    for part in parts:
        if "http" not in part:
            continue
        start = part.find("http")
        url = part[start:].strip().strip('"').strip("'")
        for sep in ['" ', "' ", " }", "}", "\\"]:
            if sep in url:
                url = url.split(sep)[0].strip()
        if not _is_image_url_good(url):
            continue
        key = url.split("?")[0].lower()
        if key in seen:
            continue
        seen.add(key)
        good.append(url)
        if len(good) >= 8:
            break
    return "|".join(good)


def _limpar_imagens_lixo_automatico(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df, 0
    out = df.copy().fillna("")
    total = 0
    for col in out.columns:
        if not _is_imagem_col(col):
            continue
        antes = out[col].astype(str).fillna("")
        depois = antes.map(_clean_image_urls_value)
        total += int((antes.str.strip() != depois.astype(str).str.strip()).sum())
        out[col] = depois
    return out.fillna(""), total


def _limpar_textos_lixo_automatico(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df, 0
    out = df.copy().fillna("")
    total = 0
    for col in out.columns:
        if _is_imagem_col(col):
            continue
        antes = out[col].astype(str).fillna("")
        depois = antes.str.replace(r"\s+", " ", regex=True).str.strip()
        depois = depois.map(lambda v: "" if ("{\"" in str(v) or "\\\"" in str(v) or str(v).lower().startswith("nan")) else v)
        total += int((antes.str.strip() != depois.astype(str).str.strip()).sum())
        out[col] = depois
    return out.fillna(""), total


def _blindar_df_capturado_site(df: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df
    df_limpo, removidos_texto = _limpar_textos_lixo_automatico(df)
    df_limpo, removidos_img = _limpar_imagens_lixo_automatico(df_limpo)
    df_limpo, removidos_gtin = _limpar_gtins_invalidos_automatico(df_limpo)
    df_limpo, removidos_ncm = _limpar_ncms_invalidos_automatico(df_limpo)
    st.session_state["stable_textos_lixo_removidos"] = removidos_texto
    st.session_state["stable_imagens_lixo_removidas"] = removidos_img
    st.session_state["stable_gtins_invalidos_removidos"] = removidos_gtin
    st.session_state["stable_ncms_invalidos_removidos"] = removidos_ncm
    return df_limpo


def _limpar_estado_dados_antigos() -> None:
    limpar_vault(prefixes=("stable_df_origem", "stable_df_export", "supplier_"))
    for key in list(st.session_state.keys()):
        if key.startswith("stable_map_") or key.startswith("supplier_"):
            st.session_state.pop(key, None)
    for key in (
        "stable_df_origem", "stable_df_export", "stable_textos_lixo_removidos",
        "stable_imagens_lixo_removidas", "stable_gtins_invalidos_removidos", "stable_ncms_invalidos_removidos",
    ):
        st.session_state.pop(key, None)


def _force(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    out = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    for c in cols:
        if c not in out.columns:
            out[c] = ""
    return out[cols].fillna("")


def _normalize_for_final(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return _force(pd.DataFrame(), cols)
    cleaned = clean_invalid_preview_mappings(df.copy().fillna(""))
    cleaned = _blindar_df_capturado_site(cleaned)
    return _force(cleaned, cols)


def _urls(raw: str) -> list[str]:
    seen, urls = set(), []
    for part in str(raw or "").replace("\r", "\n").replace(";", "\n").split("\n"):
        url = part.strip()
        if not url:
            continue
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
        if url not in seen:
            seen.add(url)
            urls.append(url)
    return urls


def _excel(df: pd.DataFrame) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="bling")
    return buf.getvalue()


def _df_chunks(df: pd.DataFrame, chunk_size: int = BLING_MAX_IMPORT_ROWS) -> list[pd.DataFrame]:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return []
    return [df.iloc[start:start + chunk_size].copy() for start in range(0, len(df), chunk_size)]


def _zip_csv_parts(df: pd.DataFrame, *, chunk_size: int = BLING_MAX_IMPORT_ROWS) -> bytes:
    buffer = BytesIO()
    parts = _df_chunks(df, chunk_size=chunk_size)
    total_parts = max(len(parts), 1)
    with ZipFile(buffer, mode="w", compression=ZIP_DEFLATED) as zip_file:
        for index, part in enumerate(parts, start=1):
            file_name = f"bling_saida_final_{len(df)}_linhas_parte_{index:02d}_de_{total_parts:02d}.csv"
            zip_file.writestr(file_name, dataframe_para_csv_bytes(part))
    return buffer.getvalue()


def _first_non_empty_value(df: pd.DataFrame, col: str) -> str:
    if not _has_df(df) or col not in df.columns:
        return ""
    try:
        for valor in df[col].astype(str).fillna(""):
            texto = str(valor or "").strip()
            if texto:
                return texto[:350]
    except Exception:
        return ""
    return ""


def _render_source_preview(df: pd.DataFrame, target: str, selected_col: str, auto_100: bool = False) -> None:
    if not selected_col:
        st.markdown("<div style='margin-top:-0.65rem;margin-bottom:0.75rem;color:#b91c1c;font-size:0.88rem;'>⚠️ Selecione a coluna correta.</div>", unsafe_allow_html=True)
        return

    analise = analyze_mapping(df, target, selected_col)
    valor = _first_non_empty_value(df, selected_col) or analise.sample_value or "sem valor preenchido"
    cor = "#047857" if analise.status == "valid" else ("#b45309" if analise.status == "warning" else "#b91c1c")
    icone = "✅" if analise.status == "valid" else ("⚠️" if analise.status == "warning" else "❌")
    texto = f"{icone} {selected_col}"
    st.markdown(
        "<div style='margin-top:-0.65rem;margin-bottom:0.75rem;line-height:1.25;'>"
        f"<div style='color:{cor};font-size:0.86rem;font-weight:700;'>{escape(texto)}</div>"
        f"<div style='color:{cor};font-size:0.84rem;font-weight:700;'>{escape(str(valor))}</div>"
        "</div>",
        unsafe_allow_html=True,
    )


def _render_downloads(out: pd.DataFrame, saida_ok: bool) -> None:
    total_linhas = len(out) if isinstance(out, pd.DataFrame) else 0
    out = _blindar_df_capturado_site(out)
    if total_linhas > BLING_MAX_IMPORT_ROWS:
        total_partes = ceil(total_linhas / BLING_MAX_IMPORT_ROWS)
        st.warning(f"O Bling aceita no máximo {BLING_MAX_IMPORT_ROWS} linhas por importação. Seu arquivo será dividido em {total_partes} partes.")
        st.download_button("📦 Baixar CSVs divididos", data=_zip_csv_parts(out), file_name=f"bling_saida_final_{total_linhas}_linhas_dividido.zip", mime="application/zip", use_container_width=True, disabled=not saida_ok)
    else:
        st.download_button(f"📥 Baixar CSV para Bling ({total_linhas} linhas)", data=dataframe_para_csv_bytes(out), file_name=f"bling_saida_final_{total_linhas}_linhas.csv", mime="text/csv", use_container_width=True, disabled=not saida_ok)
    st.download_button(f"📥 Baixar Excel de conferência ({total_linhas} linhas)", data=_excel(out), file_name=f"bling_saida_final_{total_linhas}_linhas_conferencia.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, disabled=not saida_ok)


def _col_with_data(df: pd.DataFrame, predicate) -> bool:
    return any(predicate(col) and df[col].astype(str).str.strip().ne("").any() for col in df.columns)


def _validar_saida(df: pd.DataFrame, tipo: str) -> bool:
    if not _has_df(df):
        st.error("A saída está vazia.")
        return False
    ok = True
    if tipo == "cadastro":
        if not _col_with_data(df, _is_descricao_col):
            st.error("Campo obrigatório sem dados: Descrição/Nome do produto.")
            ok = False
        if not _col_with_data(df, _is_preco_venda_col):
            st.error("Campo obrigatório sem dados: Preço unitário/preço de venda.")
            ok = False
    if tipo == "estoque":
        deposito_cols = [c for c in df.columns if _is_deposito_col(c)]
        if deposito_cols and not any(df[c].astype(str).str.strip().ne("").any() for c in deposito_cols):
            st.error("Campo obrigatório sem dados: Depósito.")
            ok = False
    return ok


def _show_line_metrics(df_origem: pd.DataFrame, df_final: pd.DataFrame | None = None) -> None:
    cols = st.columns(2)
    cols[0].metric("Linhas capturadas na origem", len(df_origem) if isinstance(df_origem, pd.DataFrame) else 0)
    cols[1].metric("Linhas no preview final", len(df_final) if isinstance(df_final, pd.DataFrame) else 0)


def _selectbox_mapping(campo: str, options: list[str], default_value: str) -> str:
    old_key = f"stable_map_{campo}"
    st.session_state.pop(old_key, None)
    key = f"stable_map_{MAPPING_UI_VERSION}_{_norm(campo).replace(' ', '_')}"
    if key in st.session_state and st.session_state[key] not in options:
        st.session_state.pop(key, None)
    index = options.index(default_value) if default_value in options else 0
    return str(st.selectbox(str(campo), options, index=index, key=key) or "").strip()


def run_stable_app() -> None:
    restaurar_chaves_df(["stable_df_origem", "stable_df_modelo", "stable_df_export"])
    st.title("🚀 IA Planilhas → Bling")
    st.caption("Fluxo corrigido: Modelo Bling → Origem/site → Mapeamento manual → Preview final → CSV.")

    with st.sidebar:
        st.caption(f"Mapping UI: {MAPPING_UI_VERSION}")
        if st.button("🧹 Reiniciar fluxo", use_container_width=True):
            limpar_vault(prefixes=("stable_", "supplier_"))
            st.rerun()

    modelo = restaurar_df("stable_df_modelo")

    with st.expander("1. Anexar modelo Bling", expanded=not _has_model_columns(modelo)):
        up_modelo = st.file_uploader("Anexar modelo Bling", type=None, key="stable_upload_modelo")
        if up_modelo is not None:
            modelo_lido = _read_bling_model_upload(up_modelo)
            if _has_model_columns(modelo_lido):
                modelo = guardar_df("stable_df_modelo", modelo_lido)
                tipo_detectado = _detect_tipo_by_model(modelo)
                st.session_state["stable_tipo"] = tipo_detectado
                st.success(f"Modelo reconhecido automaticamente: {'Atualização de estoque' if tipo_detectado == 'estoque' else 'Cadastro de produtos'}")
                st.caption(f"Modelo lido: {len(modelo.columns)} colunas. O sistema usa essa estrutura como destino.")
            else:
                st.error("Não consegui ler o modelo Bling anexado.")
                st.code("O arquivo foi reconhecido, mas não encontrei uma linha de cabeçalhos válida. Envie o modelo .xlsx exportado pelo Bling ou uma planilha com os nomes das colunas na primeira linha útil.")
        elif _has_model_columns(modelo):
            tipo_detectado = _detect_tipo_by_model(modelo)
            st.session_state["stable_tipo"] = tipo_detectado
            st.info(f"🔒 Modelo preservado e reconhecido como {'Atualização de estoque' if tipo_detectado == 'estoque' else 'Cadastro de produtos'}: {len(modelo.columns)} colunas")

    if not _has_model_columns(modelo):
        st.warning("Anexe primeiro a planilha modelo Bling. Depois disso o sistema reconhece sozinho se é cadastro ou estoque.")
        return

    tipo = str(st.session_state.get("stable_tipo") or _detect_tipo_by_model(modelo))
    cols = _cols(tipo, modelo)
    st.success(f"Operação reconhecida: {'Atualização de estoque' if tipo == 'estoque' else 'Cadastro de produtos'}")
    st.divider()
    df = restaurar_df("stable_df_origem")
    tab_file, tab_site = st.tabs(["📎 Arquivo", "🌐 Site"])

    with tab_file:
        df_up = render_supplier_upload_v2(state_key="stable_df_origem", key_prefix="supplier")
        if _has_df(df_up):
            _limpar_estado_dados_antigos()
            df = guardar_df("stable_df_origem", _blindar_df_capturado_site(df_up))

    with tab_site:
        raw = st.text_area("Links do fornecedor", key="stable_site_urls", height=120)
        links = _urls(raw)
        if st.button("Gerar base por site", disabled=not links, use_container_width=True):
            _limpar_estado_dados_antigos()
            df = executar_flash_amplo_pagina_por_pagina(links, max_products=FLASH_MAX_PRODUCTS, max_workers=12, show_progress=True)
            if _has_df(df):
                df = guardar_df("stable_df_origem", _blindar_df_capturado_site(df))
            st.rerun()

    if not _has_df(df):
        st.warning("Anexe/capture a origem para continuar.")
        return

    df = _blindar_df_capturado_site(df)
    guardar_df("stable_df_origem", df)

    _show_line_metrics(df, restaurar_df("stable_df_export"))
    st.subheader("Mapeamento manual")
    st.caption("As opções de preenchimento vêm somente do preview da origem/captura. O modelo Bling fica como destino.")

    mapping = {}
    auto_100_count = 0
    deposito_manual = ""
    for c in cols:
        options = _source_options_for_target(str(c), df, cols)
        if tipo == "estoque" and _is_deposito_col(c):
            deposito_manual = st.text_input(str(c), value=str(st.session_state.get("stable_deposito_mapeamento", "")), key="stable_deposito_mapeamento", placeholder="Ex.: Geral").strip()
            mapping[c] = ""
            continue

        auto_100 = _auto_map_100(str(c), options, df)
        if auto_100:
            auto_100_count += 1

        selecionada = _selectbox_mapping(str(c), options, auto_100)
        mapping[c] = selecionada
        _render_source_preview(df, str(c), selecionada, auto_100=bool(auto_100 and selecionada == auto_100))

    st.caption(f"Mapeamentos automáticos validados: {auto_100_count}. O restante fica vazio ou com alerta para revisão manual.")

    out = pd.DataFrame(index=df.index)
    for c in cols:
        if tipo == "estoque" and _is_deposito_col(c):
            out[c] = deposito_manual
            continue
        src = mapping.get(c, "")
        out[c] = df[src].astype(str).fillna("") if src and src in df.columns else ""
    out = _normalize_for_final(out, cols)
    out = guardar_df("stable_df_export", out)

    removidos_texto = int(st.session_state.get("stable_textos_lixo_removidos", 0) or 0)
    removidos_img = int(st.session_state.get("stable_imagens_lixo_removidas", 0) or 0)
    removidos_gtin = int(st.session_state.get("stable_gtins_invalidos_removidos", 0) or 0)
    removidos_ncm = int(st.session_state.get("stable_ncms_invalidos_removidos", 0) or 0)
    if removidos_texto:
        st.info(f"Dados antigos/lixo textual limpos automaticamente: {removidos_texto}.")
    if removidos_img:
        st.info(f"Imagens lixo removidas automaticamente: {removidos_img}. Foram mantidas apenas URLs reais de imagem de produto.")
    if removidos_gtin:
        st.info(f"GTINs inválidos limpos automaticamente: {removidos_gtin}. Campos inválidos foram deixados em branco para o Bling aceitar.")
    if removidos_ncm:
        st.info(f"NCMs inválidos limpos automaticamente: {removidos_ncm}. NCM sem 8 dígitos foi deixado em branco.")

    _show_line_metrics(df, out)
    with st.expander("Preview final", expanded=False):
        st.caption(f"Total real do arquivo final: {len(out)} linhas.")
        st.dataframe(out, use_container_width=True, hide_index=True)

    saida_ok = _validar_saida(out, tipo)
    if saida_ok:
        st.success(f"Arquivo pronto: {len(out)} linhas × {len(out.columns)} colunas")
    else:
        st.warning("Corrija o mapeamento obrigatório antes de baixar.")
    _render_downloads(out, saida_ok)
