from __future__ import annotations

import csv
from copy import copy
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any

import pandas as pd

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

SUPPORTED_TEMPLATE_EXTENSIONS = {'xlsx', 'xlsm', 'csv'}
SPREADSHEET_TEMPLATE_EXTENSIONS = {'xlsx', 'xlsm'}
TEXT_TEMPLATE_EXTENSIONS = {'csv'}
DEFAULT_HEADER_SCAN_ROWS = 80
MIN_CONTRACT_MATCH_RATIO = 1.0
DEFAULT_CSV_SEPARATOR = ';'
CSV_SEPARATOR_CANDIDATES = (';', ',', '\t', '|')


@dataclass(frozen=True)
class TemplateSheetMatch:
    sheet_name: str
    header_row: int
    positions: dict[str, int]
    missing_columns: tuple[str, ...]
    score: int
    total_columns: int

    @property
    def complete(self) -> bool:
        return self.total_columns > 0 and not self.missing_columns and len(self.positions) == self.total_columns


def _normalize(value: object) -> str:
    return ' '.join(str(value or '').replace('\ufeff', '').replace('\xa0', ' ').split()).strip().lower()


def _template_extension(file_name: str) -> str:
    name = str(file_name or '').lower().strip()
    return name.rsplit('.', 1)[-1] if '.' in name else ''


def can_export_from_template(file_name: str | None, file_bytes: bytes | None) -> bool:
    if not file_name or not file_bytes:
        return False
    extension = _template_extension(file_name)
    if extension in TEXT_TEMPLATE_EXTENSIONS:
        return True
    return extension in SPREADSHEET_TEMPLATE_EXTENSIONS and callable(load_workbook)


def output_name_for_template(file_name: str | None) -> str:
    name = str(file_name or 'modelo_preenchido.xlsx').strip() or 'modelo_preenchido.xlsx'
    if '.' not in name:
        return f'{name}_preenchido.xlsx'
    stem, ext = name.rsplit('.', 1)
    return f'{stem}_preenchido.{ext}'


def mime_for_template_output(file_name: str | None) -> str:
    extension = _template_extension(file_name)
    if extension == 'csv':
        return 'text/csv'
    if extension == 'xlsm':
        return 'application/vnd.ms-excel.sheet.macroEnabled.12'
    return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _decode_template_text(template_bytes: bytes) -> str:
    if not template_bytes:
        return ''
    for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            return template_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return template_bytes.decode('utf-8', errors='replace')


def _detect_csv_separator(template_bytes: bytes) -> str:
    text = _decode_template_text(template_bytes)
    first_lines = [line for line in text.splitlines()[:10] if line.strip()]
    if not first_lines:
        return DEFAULT_CSV_SEPARATOR

    try:
        dialect = csv.Sniffer().sniff('\n'.join(first_lines), delimiters=';,\t|')
        if dialect.delimiter in CSV_SEPARATOR_CANDIDATES:
            return dialect.delimiter
    except Exception:
        pass

    header = first_lines[0]
    counts = {separator: header.count(separator) for separator in CSV_SEPARATOR_CANDIDATES}
    best = max(counts, key=counts.get)
    return best if counts.get(best, 0) > 0 else DEFAULT_CSV_SEPARATOR


def build_csv_template_download_bytes(*, template_bytes: bytes, df: pd.DataFrame) -> bytes:
    """Gera CSV respeitando o contrato e o separador do modelo CSV anexado."""
    if not isinstance(df, pd.DataFrame):
        raise ValueError('DataFrame final inválido.')
    separator = _detect_csv_separator(template_bytes)
    safe_df = df.copy().fillna('')
    output = StringIO()
    safe_df.to_csv(output, sep=separator, index=False, lineterminator='\n')
    return output.getvalue().encode('utf-8-sig')


def _build_column_positions(ws: Any, header_row: int, columns: list[str]) -> dict[str, int]:
    by_normalized_header: dict[str, int] = {}
    for col_index in range(1, (ws.max_column or 1) + 1):
        value = _normalize(ws.cell(row=header_row, column=col_index).value)
        if value and value not in by_normalized_header:
            by_normalized_header[value] = col_index

    positions: dict[str, int] = {}
    for column in columns:
        key = _normalize(column)
        if key in by_normalized_header:
            positions[column] = by_normalized_header[key]
    return positions


def _match_sheet_header(ws: Any, columns: list[str]) -> TemplateSheetMatch:
    wanted = [_normalize(column) for column in columns if _normalize(column)]
    total = len(wanted)
    if total <= 0:
        return TemplateSheetMatch(str(ws.title), 1, {}, tuple(columns), 0, 0)

    best_row = 1
    best_positions: dict[str, int] = {}
    best_score = -1
    max_scan = min(max(ws.max_row or 1, 1), DEFAULT_HEADER_SCAN_ROWS)

    for row_index in range(1, max_scan + 1):
        positions = _build_column_positions(ws, row_index, columns)
        score = len(positions)
        if score > best_score:
            best_score = score
            best_row = row_index
            best_positions = positions
        if score == total:
            break

    missing = tuple(column for column in columns if column not in best_positions)
    return TemplateSheetMatch(str(ws.title), best_row, best_positions, missing, max(best_score, 0), total)


def find_best_template_sheet(workbook: Any, columns: list[str]) -> tuple[Any, TemplateSheetMatch]:
    matches: list[tuple[Any, TemplateSheetMatch]] = []
    for worksheet in workbook.worksheets:
        matches.append((worksheet, _match_sheet_header(worksheet, columns)))

    if not matches:
        raise ValueError('O arquivo modelo não possui abas disponíveis.')

    matches.sort(key=lambda item: (item[1].score, -item[0].max_row, -item[0].max_column), reverse=True)
    worksheet, match = matches[0]

    required_matches = int(match.total_columns * MIN_CONTRACT_MATCH_RATIO)
    if not match.complete or match.score < required_matches:
        missing_preview = ', '.join(match.missing_columns[:30])
        if len(match.missing_columns) > 30:
            missing_preview += '...'
        raise ValueError(
            'Contrato rígido bloqueou o download: não encontrei 100% das colunas do modelo original. '
            f'Aba mais próxima: {match.sheet_name}. Encontradas: {match.score}/{match.total_columns}. '
            f'Faltando: {missing_preview}'
        )

    return worksheet, match


def _copy_row_style(ws: Any, source_row: int, target_row: int) -> None:
    if source_row <= 0 or target_row <= 0 or source_row == target_row:
        return

    try:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden
        ws.row_dimensions[target_row].outlineLevel = ws.row_dimensions[source_row].outlineLevel
    except Exception:
        pass

    for col_index in range(1, (ws.max_column or 1) + 1):
        source = ws.cell(row=source_row, column=col_index)
        target = ws.cell(row=target_row, column=col_index)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def _write_dataframe_to_sheet(ws: Any, df: pd.DataFrame, match: TemplateSheetMatch) -> None:
    first_data_row = match.header_row + 1
    template_data_row = first_data_row if ws.max_row >= first_data_row else match.header_row
    total_rows = len(df)
    last_needed_row = first_data_row + max(total_rows - 1, 0)
    max_existing_row = max(ws.max_row or first_data_row, last_needed_row)

    for row_index in range(first_data_row, last_needed_row + 1):
        if row_index > (ws.max_row or 0):
            _copy_row_style(ws, template_data_row, row_index)

    safe_df = df.copy().fillna('')
    for row_offset, (_, row) in enumerate(safe_df.iterrows()):
        excel_row = first_data_row + row_offset
        if excel_row > (ws.max_row or 0):
            _copy_row_style(ws, template_data_row, excel_row)
        for column, col_index in match.positions.items():
            ws.cell(row=excel_row, column=col_index).value = '' if pd.isna(row[column]) else row[column]

    for excel_row in range(first_data_row + total_rows, max_existing_row + 1):
        for col_index in match.positions.values():
            ws.cell(row=excel_row, column=col_index).value = None


def build_spreadsheet_template_download_bytes(
    *,
    template_bytes: bytes,
    template_name: str,
    df: pd.DataFrame,
) -> bytes:
    if not callable(load_workbook):
        raise ValueError('Leitor de planilha indisponível para exportação fiel.')
    if not isinstance(df, pd.DataFrame):
        raise ValueError('DataFrame final inválido.')

    safe_df = df.copy().fillna('')
    columns = [str(column) for column in safe_df.columns]
    keep_vba = _template_extension(template_name) == 'xlsm'
    workbook = load_workbook(BytesIO(template_bytes), keep_vba=keep_vba)
    worksheet, match = find_best_template_sheet(workbook, columns)
    _write_dataframe_to_sheet(worksheet, safe_df, match)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_template_download_bytes(
    *,
    template_bytes: bytes,
    template_name: str,
    df: pd.DataFrame,
) -> bytes:
    extension = _template_extension(template_name)
    if extension == 'csv':
        return build_csv_template_download_bytes(template_bytes=template_bytes, df=df)
    if extension in SPREADSHEET_TEMPLATE_EXTENSIONS:
        return build_spreadsheet_template_download_bytes(template_bytes=template_bytes, template_name=template_name, df=df)
    raise ValueError('Formato de modelo ainda não suportado para download 100% fiel. Use CSV, XLSX ou XLSM.')


__all__ = [
    'CSV_SEPARATOR_CANDIDATES',
    'DEFAULT_CSV_SEPARATOR',
    'SUPPORTED_TEMPLATE_EXTENSIONS',
    'SPREADSHEET_TEMPLATE_EXTENSIONS',
    'TEXT_TEMPLATE_EXTENSIONS',
    'TemplateSheetMatch',
    'build_csv_template_download_bytes',
    'build_spreadsheet_template_download_bytes',
    'build_template_download_bytes',
    'can_export_from_template',
    'find_best_template_sheet',
    'mime_for_template_output',
    'output_name_for_template',
]