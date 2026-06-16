from __future__ import annotations

import csv
from copy import copy
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Sequence

import pandas as pd

CSV_MIME = 'text/csv; charset=utf-8'
XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
XLSM_MIME = 'application/vnd.ms-excel.sheet.macroEnabled.12'
PRESERVABLE_EXCEL_SUFFIXES = {'.xlsx', '.xlsm'}
PRESERVABLE_TEXT_SUFFIXES = {'.csv'}
PRESERVABLE_SUFFIXES = PRESERVABLE_EXCEL_SUFFIXES | PRESERVABLE_TEXT_SUFFIXES


@dataclass(frozen=True)
class PreservedTemplateExportResult:
    file_name: str
    file_bytes: bytes
    mime: str
    format: str


def _clean(value: object) -> str:
    if value is None:
        return ''
    text = str(value).replace('\ufeff', '').replace('\x00', '').replace('\xa0', ' ')
    return ' '.join(text.split()).strip()


def _norm(value: object) -> str:
    return _clean(value).casefold()


def template_can_be_preserved(file_name: str | None) -> bool:
    return Path(str(file_name or '')).suffix.lower() in PRESERVABLE_SUFFIXES


def template_download_mime(file_name: str | None) -> str:
    suffix = Path(str(file_name or '')).suffix.lower()
    if suffix == '.xlsx':
        return XLSX_MIME
    if suffix == '.xlsm':
        return XLSM_MIME
    return CSV_MIME


def _dataframe_columns(df: pd.DataFrame | None) -> list[str]:
    if not isinstance(df, pd.DataFrame):
        return []
    return [_clean(column) for column in df.columns]


def _csv_separator_from_text(text: str) -> str:
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=',;\t|').delimiter
    except Exception:
        candidates = [',', ';', '\t', '|']
        return max(candidates, key=lambda sep: sample.count(sep))


def _decode_text(data: bytes) -> str:
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin1'):
        try:
            return data.decode(encoding)
        except Exception:
            continue
    return data.decode('utf-8', errors='ignore')


def _csv_header_columns(template_bytes: bytes) -> list[str]:
    text = _decode_text(template_bytes)
    first_line = next((line for line in text.splitlines() if line.strip()), '')
    if not first_line:
        return []
    sep = _csv_separator_from_text(text)
    return [_clean(part) for part in next(csv.reader([first_line], delimiter=sep)) if _clean(part)]


def _excel_header_columns(template_name: str, template_bytes: bytes) -> list[str]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    workbook = None
    try:
        keep_vba = Path(template_name).suffix.lower() == '.xlsm'
        workbook = load_workbook(BytesIO(template_bytes), read_only=True, data_only=False, keep_vba=keep_vba)
        best: list[str] = []
        for sheet in workbook.worksheets:
            max_rows = min(int(sheet.max_row or 1), 30)
            max_cols = int(sheet.max_column or 0)
            for row in sheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
                columns = [_clean(value) for value in row if _clean(value)]
                if len(columns) > len(best):
                    best = columns
        return best
    except Exception:
        return []
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass


def template_contract_columns(template_name: str | None, template_bytes: bytes | bytearray | None) -> list[str]:
    if not template_name or not template_bytes:
        return []
    suffix = Path(str(template_name)).suffix.lower()
    data = bytes(template_bytes)
    if suffix == '.csv':
        return _csv_header_columns(data)
    if suffix in PRESERVABLE_EXCEL_SUFFIXES or suffix in {'.xls', '.xlsb'}:
        return _excel_header_columns(str(template_name), data)
    return []


def _header_match(row_values: Sequence[object], wanted_columns: Sequence[str]) -> tuple[int, dict[str, int]]:
    wanted_set = {_norm(column) for column in wanted_columns if _norm(column)}
    mapping: dict[str, int] = {}
    score = 0
    for index, value in enumerate(row_values, start=1):
        key = _norm(value)
        if key and key in wanted_set and key not in mapping:
            mapping[key] = index
            score += 1
    return score, mapping


def _find_header_location(workbook, wanted_columns: list[str]):
    wanted_count = len([column for column in wanted_columns if _clean(column)])
    best_score = -1
    best = None
    for sheet in workbook.worksheets:
        max_rows = min(int(sheet.max_row or 1), 30)
        max_cols = max(int(sheet.max_column or 0), wanted_count)
        for row_index in range(1, max_rows + 1):
            values = [sheet.cell(row=row_index, column=col_index).value for col_index in range(1, max_cols + 1)]
            score, mapping = _header_match(values, wanted_columns)
            if score > best_score:
                best_score = score
                best = (sheet, row_index, mapping)
            if wanted_count and score >= wanted_count:
                return sheet, row_index, mapping
    if best is not None and best_score > 0:
        return best
    return workbook.worksheets[0], 1, {}


def _copy_cell_style(source_cell, target_cell) -> None:
    try:
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)
    except Exception:
        pass


def _excel_bytes_from_template(template_name: str, template_bytes: bytes, df: pd.DataFrame) -> bytes:
    from openpyxl import load_workbook

    keep_vba = Path(template_name).suffix.lower() == '.xlsm'
    workbook = load_workbook(BytesIO(template_bytes), keep_vba=keep_vba, data_only=False)
    columns = _dataframe_columns(df)
    sheet, header_row, header_mapping = _find_header_location(workbook, columns)
    normalized_mapping = dict(header_mapping or {})
    if not normalized_mapping:
        normalized_mapping = {_norm(column): index for index, column in enumerate(columns, start=1)}
    target_columns = {column: normalized_mapping.get(_norm(column), index) for index, column in enumerate(columns, start=1)}

    last_row_to_clear = max(int(sheet.max_row or header_row), header_row + len(df))
    for row_index in range(header_row + 1, last_row_to_clear + 1):
        for target_col in target_columns.values():
            sheet.cell(row=row_index, column=target_col).value = None

    style_row = header_row + 1 if int(sheet.max_row or header_row) >= header_row + 1 else header_row
    for row_offset, (_, series) in enumerate(df.fillna('').astype(str).iterrows(), start=1):
        row_index = header_row + row_offset
        for column in columns:
            target_col = target_columns[column]
            target_cell = sheet.cell(row=row_index, column=target_col)
            if row_index > style_row:
                _copy_cell_style(sheet.cell(row=style_row, column=target_col), target_cell)
            target_cell.value = _clean(series.get(column))

    output = BytesIO()
    workbook.save(output)
    try:
        workbook.close()
    except Exception:
        pass
    return output.getvalue()


def _csv_bytes_with_template_name(template_bytes: bytes, df: pd.DataFrame) -> bytes:
    text = _decode_text(template_bytes)
    sep = _csv_separator_from_text(text) if text.strip() else ';'
    buffer = StringIO()
    df.fillna('').astype(str).to_csv(buffer, sep=sep, index=False, encoding='utf-8-sig', quoting=csv.QUOTE_MINIMAL, lineterminator='\n')
    return buffer.getvalue().encode('utf-8-sig')


def build_preserved_template_export(template_name: str | None, template_bytes: bytes | bytearray | None, df: pd.DataFrame | None) -> PreservedTemplateExportResult | None:
    if not template_name or not template_bytes or not isinstance(df, pd.DataFrame):
        return None
    suffix = Path(str(template_name)).suffix.lower()
    if suffix not in PRESERVABLE_SUFFIXES:
        return None
    data = bytes(template_bytes)
    safe_df = df.copy().fillna('')
    if suffix in PRESERVABLE_EXCEL_SUFFIXES:
        file_bytes = _excel_bytes_from_template(str(template_name), data, safe_df)
        return PreservedTemplateExportResult(str(template_name), file_bytes, template_download_mime(str(template_name)), suffix.lstrip('.'))
    if suffix == '.csv':
        file_bytes = _csv_bytes_with_template_name(data, safe_df)
        return PreservedTemplateExportResult(str(template_name), file_bytes, CSV_MIME, 'csv')
    return None


__all__ = [
    'CSV_MIME',
    'PRESERVABLE_EXCEL_SUFFIXES',
    'PreservedTemplateExportResult',
    'XLSM_MIME',
    'XLSX_MIME',
    'build_preserved_template_export',
    'template_can_be_preserved',
    'template_contract_columns',
    'template_download_mime',
]
