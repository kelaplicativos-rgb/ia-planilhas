from __future__ import annotations

import csv
import re
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any

import pandas as pd

from bling_app_zero.core.html_product_extractor import (
    clean_columns as _clean_html_columns,
    clean_text as _html_clean,
    decode_html_bytes,
    read_html_product_bytes,
    read_mhtml_product_bytes,
)

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

SPREADSHEET_EXTENSIONS = ('.xlsx', '.xls', '.xlsm', '.xlsb')
TEXT_EXTENSIONS = ('.txt', '.tsv')
HTML_EXTENSIONS = ('.html', '.htm')
MHTML_EXTENSIONS = ('.mht', '.mhtml')
ZIP_EXTENSIONS = ('.zip',)
SUPPORTED_SUPPLIER_EXTENSIONS = SPREADSHEET_EXTENSIONS + ('.csv', '.xml', '.pdf') + TEXT_EXTENSIONS + HTML_EXTENSIONS + MHTML_EXTENSIONS + ZIP_EXTENSIONS

NFE_ITEM_TAGS = {'det'}
NFE_PRODUCT_TAGS = {'prod'}
COMMON_ITEM_TAGS = {'item', 'produto', 'product', 'det', 'row', 'linha'}
EXCEL_HEADER_SCAN_ROWS = 80
EXCEL_DATA_SCAN_ROWS = 25
HEADER_KEYWORDS = {
    'id', 'codigo', 'código', 'cod', 'sku', 'ref', 'referencia', 'referência', 'ean', 'gtin',
    'produto', 'produtos', 'nome', 'titulo', 'título', 'descricao', 'descrição', 'categoria',
    'marca', 'modelo', 'ncm', 'origem', 'preco', 'preço', 'valor', 'custo', 'estoque',
    'quantidade', 'qtde', 'deposito', 'depósito', 'unidade', 'peso', 'altura', 'largura',
    'comprimento', 'imagem', 'imagens', 'url', 'link', 'atributo', 'variacao', 'variação',
    'status', 'situacao', 'situação', 'canal', 'venda', 'loja', 'observacao', 'observação',
}
INSTRUCTION_SHEET_HINTS = ('instr', 'instru', 'leia', 'ajuda', 'help', 'readme', 'manual')


@dataclass(frozen=True)
class ExcelHeaderCandidate:
    sheet_name: str
    sheet_index: int
    header_row: int
    max_column: int
    columns: tuple[str, ...]
    positions: tuple[int, ...]
    score: float
    data_rows_below: int


def _decode_bytes(data: bytes) -> str:
    return decode_html_bytes(data)


def _detect_separator(text: str) -> str:
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        return dialect.delimiter
    except Exception:
        candidates = [',', ';', '\t', '|']
        return max(candidates, key=lambda sep: sample.count(sep))


def _clean(value: object) -> str:
    return _html_clean(value)


def _norm(value: object) -> str:
    text = _clean(value).casefold()
    text = re.sub(r'[^0-9a-záàâãéêíóôõúçñ]+', ' ', text)
    return ' '.join(text.split()).strip()


def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    return _clean_html_columns(df)


def _best_frame(frames: list[pd.DataFrame]) -> pd.DataFrame:
    """Escolhe a melhor tabela preservando planilhas modelo sem linhas."""
    structured: list[pd.DataFrame] = []
    header_only: list[pd.DataFrame] = []

    for frame in frames:
        if not isinstance(frame, pd.DataFrame) or len(frame.columns) == 0:
            continue
        cleaned = _clean_columns(frame).fillna('').astype(str).reset_index(drop=True)
        if cleaned.empty:
            header_only.append(cleaned)
        else:
            structured.append(cleaned)

    if structured:
        structured.sort(key=lambda df: len(df) * max(1, len(df.columns)), reverse=True)
        return structured[0]

    if header_only:
        header_only.sort(key=lambda df: len(df.columns), reverse=True)
        return header_only[0]

    return pd.DataFrame()


def _read_csv_bytes(data: bytes) -> pd.DataFrame:
    text = _decode_bytes(data)
    sep = _detect_separator(text)
    return _clean_columns(pd.read_csv(StringIO(text), sep=sep, dtype=str).fillna(''))


def _read_text_bytes(data: bytes) -> pd.DataFrame:
    text = _decode_bytes(data)
    if '\t' in text[:8192]:
        return _clean_columns(pd.read_csv(StringIO(text), sep='\t', dtype=str).fillna(''))
    rows = [[_clean(part) for part in re.split(r'\s{2,}|;|\|', line) if _clean(part)] for line in text.splitlines() if _clean(line)]
    rows = [row for row in rows if row]
    if len(rows) < 2:
        return pd.DataFrame([{'Arquivo TXT': 'texto', 'Conteúdo': text[:20000]}]) if text.strip() else pd.DataFrame()
    width = max(len(row) for row in rows)
    normalized = [row + [''] * (width - len(row)) for row in rows]
    columns = [_clean(value) or f'Coluna {idx + 1}' for idx, value in enumerate(normalized[0])]
    return _clean_columns(pd.DataFrame(normalized[1:], columns=columns))


def _cell_value(cell: Any) -> str:
    return _clean(getattr(cell, 'value', cell))


def _row_values(sheet: Any, row_index: int, max_col: int) -> list[str]:
    return [_cell_value(sheet.cell(row=row_index, column=col_index)) for col_index in range(1, max_col + 1)]


def _dedupe_columns(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for idx, column in enumerate(columns, start=1):
        base = _clean(column) or f'Coluna {idx}'
        key = base
        counter = seen.get(base, 0)
        if counter:
            key = f'{base}_{counter + 1}'
        seen[base] = counter + 1
        out.append(key)
    return out


def _keyword_hits(values: list[str]) -> int:
    hits = 0
    for value in values:
        norm = _norm(value)
        words = set(norm.split())
        if words & HEADER_KEYWORDS:
            hits += 1
        elif any(keyword in norm for keyword in HEADER_KEYWORDS if len(keyword) >= 4):
            hits += 1
    return hits


def _looks_like_instruction_row(values: list[str]) -> bool:
    non_empty = [value for value in values if _clean(value)]
    if not non_empty:
        return True
    if len(non_empty) <= 1 and len(non_empty[0]) > 40:
        return True
    long_cells = sum(1 for value in non_empty if len(value) > 80)
    return long_cells >= max(2, len(non_empty) // 2)


def _data_rows_below(sheet: Any, header_row: int, positions: list[int]) -> int:
    if not positions:
        return 0
    max_row = int(sheet.max_row or header_row)
    limit = min(max_row, header_row + EXCEL_DATA_SCAN_ROWS)
    count = 0
    for row_index in range(header_row + 1, limit + 1):
        row_values = [_cell_value(sheet.cell(row=row_index, column=col_index)) for col_index in positions]
        filled = sum(1 for value in row_values if value)
        if filled >= 1:
            count += 1
    return count


def _score_header_candidate(sheet: Any, sheet_index: int, row_index: int, max_col: int) -> ExcelHeaderCandidate | None:
    values = _row_values(sheet, row_index, max_col)
    positions = [idx + 1 for idx, value in enumerate(values) if _clean(value)]
    raw_columns = [_clean(values[idx - 1]) for idx in positions]
    if len(raw_columns) < 2:
        return None
    if _looks_like_instruction_row(raw_columns):
        return None

    columns = tuple(_dedupe_columns(raw_columns))
    normalized = [_norm(value) for value in raw_columns]
    unique_count = len(set(value for value in normalized if value))
    keyword_hits = _keyword_hits(raw_columns)
    data_rows = _data_rows_below(sheet, row_index, positions)
    avg_len = sum(len(value) for value in raw_columns) / max(1, len(raw_columns))
    sheet_name = str(getattr(sheet, 'title', '') or '')
    sheet_hint_penalty = 18 if any(hint in _norm(sheet_name) for hint in INSTRUCTION_SHEET_HINTS) else 0
    long_penalty = 12 if avg_len > 45 else 0
    score = (
        len(raw_columns) * 10
        + unique_count * 4
        + keyword_hits * 16
        + min(data_rows, 8) * 5
        - sheet_hint_penalty
        - long_penalty
        - row_index * 0.15
    )
    return ExcelHeaderCandidate(
        sheet_name=sheet_name,
        sheet_index=sheet_index,
        header_row=row_index,
        max_column=max_col,
        columns=columns,
        positions=tuple(positions),
        score=score,
        data_rows_below=data_rows,
    )


def _find_best_excel_header(workbook: Any) -> ExcelHeaderCandidate | None:
    candidates: list[ExcelHeaderCandidate] = []
    for sheet_index, sheet in enumerate(workbook.worksheets):
        max_row = int(sheet.max_row or 0)
        max_col = int(sheet.max_column or 0)
        if max_row <= 0 or max_col <= 0:
            continue
        scan_rows = min(max_row, EXCEL_HEADER_SCAN_ROWS)
        for row_index in range(1, scan_rows + 1):
            candidate = _score_header_candidate(sheet, sheet_index, row_index, max_col)
            if candidate is not None:
                candidates.append(candidate)
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item.score, item.data_rows_below, len(item.columns), -item.header_row), reverse=True)
    return candidates[0]


def _frame_from_excel_candidate(workbook: Any, candidate: ExcelHeaderCandidate) -> pd.DataFrame:
    sheet = workbook.worksheets[candidate.sheet_index]
    rows: list[list[str]] = []
    max_row = int(sheet.max_row or candidate.header_row)
    for row_index in range(candidate.header_row + 1, max_row + 1):
        row = [_cell_value(sheet.cell(row=row_index, column=col_index)) for col_index in candidate.positions]
        if any(_clean(value) for value in row):
            rows.append(row)
    return _clean_columns(pd.DataFrame(rows, columns=list(candidate.columns)).fillna(''))


def _read_excel_with_header_detection(data: bytes, file_name: str) -> pd.DataFrame:
    if load_workbook is None:
        return pd.DataFrame()
    workbook = None
    try:
        keep_vba = str(file_name or '').lower().endswith('.xlsm')
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=False, keep_vba=keep_vba)
        candidate = _find_best_excel_header(workbook)
        if candidate is None:
            return pd.DataFrame()
        return _frame_from_excel_candidate(workbook, candidate)
    except Exception:
        return pd.DataFrame()
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass


def _read_excel_bytes(data: bytes, file_name: str) -> pd.DataFrame:
    detected = _read_excel_with_header_detection(data, file_name)
    if isinstance(detected, pd.DataFrame) and len(detected.columns) > 0:
        return detected

    buffer = BytesIO(data)
    try:
        sheets = pd.read_excel(buffer, sheet_name=None, dtype=str).values()
        return _best_frame([sheet.fillna('') for sheet in sheets])
    except Exception:
        buffer.seek(0)
        df = pd.read_excel(buffer, dtype=str).fillna('')
        return _clean_columns(df)


def _read_html_bytes(data: bytes) -> pd.DataFrame:
    """Leitor universal de HTML.

    A saída fica mais rica para o mapeamento, mas o download continua fiel ao
    modelo anexado na primeira etapa. Este leitor apenas cria aliases úteis
    como Descrição Produto/Título/Nome/Produto para aumentar a chance de
    preencher campos de nome quando o fornecedor usa tabelas ou cards.
    """
    return read_html_product_bytes(data)


def _read_mhtml_bytes(data: bytes) -> pd.DataFrame:
    """Leitor universal de MHTML/MHT.

    Usa o extrator modular para encontrar tabelas, cards e títulos salvos no
    arquivo do navegador, sem distinguir estoque/cadastro. A distinção continua
    sendo exclusivamente o contrato/modelo anexado pelo usuário.
    """
    return read_mhtml_product_bytes(data)


def _xml_tag_name(tag: str) -> str:
    return str(tag or '').split('}', 1)[-1].strip()


def _xml_child_texts(element: ET.Element, prefix: str = '') -> dict[str, str]:
    data: dict[str, str] = {}
    for child in list(element):
        name = _xml_tag_name(child.tag)
        key = f'{prefix}{name}' if not prefix else f'{prefix}.{name}'
        children = list(child)
        text = _clean(child.text)
        if children:
            data.update(_xml_child_texts(child, key))
        elif text:
            data[key] = text
        for attr, value in child.attrib.items():
            data[f'{key}@{attr}'] = _clean(value)
    return data


def _nfe_item_to_row(det: ET.Element) -> dict[str, str]:
    row: dict[str, str] = {}
    n_item = det.attrib.get('nItem')
    if n_item:
        row['nItem'] = _clean(n_item)
    for child in list(det):
        child_name = _xml_tag_name(child.tag)
        if child_name in NFE_PRODUCT_TAGS:
            for product_child in list(child):
                row[_xml_tag_name(product_child.tag)] = _clean(product_child.text)
        else:
            row.update(_xml_child_texts(child, child_name))
    return row


def _read_xml_bytes(data: bytes, file_name: str = 'xml') -> pd.DataFrame:
    text = _decode_bytes(data)
    try:
        root = ET.fromstring(text.encode('utf-8'))
    except Exception:
        try:
            root = ET.fromstring(text)
        except Exception:
            return pd.DataFrame([{'Arquivo XML': file_name, 'Conteúdo XML': text[:50000]}])

    rows: list[dict[str, str]] = []
    for element in root.iter():
        name = _xml_tag_name(element.tag).lower()
        if name in NFE_ITEM_TAGS:
            row = _nfe_item_to_row(element)
            if row:
                rows.append(row)

    if not rows:
        for element in root.iter():
            name = _xml_tag_name(element.tag).lower()
            if name in COMMON_ITEM_TAGS:
                row = _xml_child_texts(element)
                if len(row) >= 2:
                    rows.append(row)

    if not rows:
        flat = _xml_child_texts(root)
        if flat:
            rows.append(flat)

    if not rows:
        return pd.DataFrame([{'Arquivo XML': file_name, 'Conteúdo XML': text[:50000]}])

    return _clean_columns(pd.DataFrame(rows).fillna(''))


def _split_pdf_table_line(line: str) -> list[str]:
    clean = _clean(line)
    if not clean:
        return []
    parts = [part.strip() for part in re.split(r'\s{2,}|\t|;|\|', clean) if part.strip()]
    return parts


def _read_pdf_bytes(data: bytes, file_name: str = 'pdf') -> pd.DataFrame:
    if PdfReader is None:
        return pd.DataFrame([{'Arquivo PDF': file_name, 'Status': 'Leitor de PDF indisponível no ambiente.'}])

    try:
        reader = PdfReader(BytesIO(data))
        page_texts = [page.extract_text() or '' for page in reader.pages]
        text = '\n'.join(page_texts)
    except Exception as exc:
        return pd.DataFrame([{'Arquivo PDF': file_name, 'Status': f'Não foi possível ler o PDF: {exc}'}])

    lines = [_clean(line) for line in text.splitlines() if _clean(line)]
    table_rows = [_split_pdf_table_line(line) for line in lines]
    table_rows = [row for row in table_rows if len(row) >= 2]

    if len(table_rows) >= 2:
        width = max(len(row) for row in table_rows)
        normalized = [row + [''] * (width - len(row)) for row in table_rows]
        header_index = 0
        for idx, row in enumerate(normalized[:10]):
            row_text = ' '.join(row).lower()
            if any(term in row_text for term in ('codigo', 'código', 'sku', 'produto', 'descricao', 'descrição', 'preco', 'preço', 'qtde', 'quantidade')):
                header_index = idx
                break
        columns = [_clean(value) or f'Coluna {idx + 1}' for idx, value in enumerate(normalized[header_index])]
        data_rows = normalized[header_index + 1 :]
        if data_rows:
            return _clean_columns(pd.DataFrame(data_rows, columns=columns))

    if text.strip():
        return pd.DataFrame([{'Arquivo PDF': file_name, 'Texto extraído': text[:50000]}])
    return pd.DataFrame([{'Arquivo PDF': file_name, 'Status': 'PDF sem texto extraível. Pode ser imagem/scan.'}])


def _read_inner_file_bytes(data: bytes, file_name: str) -> pd.DataFrame:
    lower = str(file_name or '').lower()
    if lower.endswith('.csv'):
        return _read_csv_bytes(data)
    if lower.endswith(TEXT_EXTENSIONS):
        return _read_text_bytes(data)
    if lower.endswith(SPREADSHEET_EXTENSIONS):
        return _read_excel_bytes(data, file_name)
    if lower.endswith(HTML_EXTENSIONS):
        return _read_html_bytes(data)
    if lower.endswith(MHTML_EXTENSIONS):
        return _read_mhtml_bytes(data)
    if lower.endswith('.xml'):
        return _read_xml_bytes(data, file_name)
    return pd.DataFrame()


def _read_zip_bytes(data: bytes, file_name: str = 'arquivo.zip') -> pd.DataFrame:
    """Lê ZIP quando ele contém CSV/XLSX interno."""
    frames: list[pd.DataFrame] = []
    try:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                inner_name = str(info.filename or '')
                lower = inner_name.lower()
                if not lower.endswith(('.csv', '.xlsx', '.xls', '.xlsm', '.xlsb', '.txt', '.tsv', '.html', '.htm', '.mht', '.mhtml', '.xml')):
                    continue
                try:
                    frame = _read_inner_file_bytes(archive.read(info), inner_name)
                except Exception:
                    frame = pd.DataFrame()
                if isinstance(frame, pd.DataFrame) and len(frame.columns) > 0:
                    frames.append(frame)
    except Exception:
        return pd.DataFrame()
    return _best_frame(frames)


def read_uploaded_file(uploaded_file: Any) -> pd.DataFrame:
    if uploaded_file is None:
        return pd.DataFrame()

    name_original = str(getattr(uploaded_file, 'name', '') or 'arquivo')
    name = name_original.lower()
    data = uploaded_file.getvalue()

    if name.endswith('.csv'):
        return _read_csv_bytes(data)
    if name.endswith(TEXT_EXTENSIONS):
        return _read_text_bytes(data)
    if name.endswith(SPREADSHEET_EXTENSIONS):
        return _read_excel_bytes(data, name_original)
    if name.endswith(HTML_EXTENSIONS):
        return _read_html_bytes(data)
    if name.endswith(MHTML_EXTENSIONS):
        return _read_mhtml_bytes(data)
    if name.endswith('.xml'):
        return _read_xml_bytes(data, name_original)
    if name.endswith('.pdf'):
        return _read_pdf_bytes(data, name_original)
    if name.endswith(ZIP_EXTENSIONS):
        return _read_zip_bytes(data, name_original)
    return pd.DataFrame()


def ensure_dataframe(value: Any) -> pd.DataFrame:
    if isinstance(value, pd.DataFrame):
        return value.copy().fillna('')
    return pd.DataFrame()


__all__ = [
    'SUPPORTED_SUPPLIER_EXTENSIONS',
    'ensure_dataframe',
    'read_uploaded_file',
]
