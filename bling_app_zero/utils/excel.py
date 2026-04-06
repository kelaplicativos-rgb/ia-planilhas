from __future__ import annotations

import csv
import re
import unicodedata
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl.styles import numbers


ENCODINGS_CSV = ("utf-8", "utf-8-sig", "latin1", "cp1252")
SEPARADORES_CSV = (",", ";", "\t", "|")


def normalizar_colunas(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]
    return df


def limpar_valores_vazios(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    df = df.copy()
    df = df.replace({pd.NA: "", "nan": "", "NaN": "", "None": ""})
    return df.fillna("")


def _ler_csv_bytes(conteudo: bytes) -> pd.DataFrame:
    ultimo_erro: Exception | None = None

    for encoding in ENCODINGS_CSV:
        try:
            texto = conteudo.decode(encoding)
        except Exception as e:
            ultimo_erro = e
            continue

        try:
            amostra = texto[:4096]
            try:
                dialect = csv.Sniffer().sniff(amostra, delimiters=";,\t|")
                candidatos = [dialect.delimiter] + [
                    s for s in SEPARADORES_CSV if s != dialect.delimiter
                ]
            except Exception:
                candidatos = list(SEPARADORES_CSV)

            for sep in candidatos:
                try:
                    df = pd.read_csv(
                        BytesIO(conteudo),
                        sep=sep,
                        encoding=encoding,
                        dtype=str,
                        keep_default_na=False,
                    )
                    if df is not None and len(df.columns) > 0:
                        return df
                except Exception as e:
                    ultimo_erro = e
                    continue
        except Exception as e:
            ultimo_erro = e
            continue

    if ultimo_erro is not None:
        raise ultimo_erro
    raise ValueError("Não foi possível ler o CSV informado.")


def ler_planilha(arquivo) -> pd.DataFrame:
    if arquivo is None:
        return pd.DataFrame()

    nome = str(getattr(arquivo, "name", "")).lower().strip()

    if hasattr(arquivo, "seek"):
        arquivo.seek(0)
    conteudo = arquivo.read()
    if hasattr(arquivo, "seek"):
        arquivo.seek(0)

    if not conteudo:
        return pd.DataFrame()

    if nome.endswith(".csv"):
        df = _ler_csv_bytes(conteudo)
    elif nome.endswith((".xlsx", ".xls")):
        df = pd.read_excel(BytesIO(conteudo), dtype=str)
    else:
        try:
            df = pd.read_excel(BytesIO(conteudo), dtype=str)
        except Exception:
            df = _ler_csv_bytes(conteudo)

    df = normalizar_colunas(df)
    df = limpar_valores_vazios(df)
    return df


def gerar_preview(df: pd.DataFrame, linhas: int = 5) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    linhas = max(1, int(linhas or 1))
    return df.head(linhas).copy()


def bloco_toggle(label: str, key: str) -> bool:
    return st.checkbox(label, value=False, key=f"toggle_{key}")


def _normalizar_texto(texto: str) -> str:
    texto = str(texto or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ascii", "ignore").decode("utf-8")
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _coluna_parece_codigo(header: str) -> bool:
    nome = _normalizar_texto(header)
    termos = {
        "codigo",
        "codigo produto",
        "id produto",
        "sku",
        "gtin",
        "ean",
        "codigo de barras",
        "ncm",
        "cean",
        "ceantrib",
        "deposito",
        "deposito obrigatorio",
        "deposito id",
        "data",
    }
    if nome in termos:
        return True

    return any(
        termo in nome
        for termo in [
            "codigo",
            "sku",
            "gtin",
            "ean",
            "barras",
            "ncm",
            "deposito",
            "data",
        ]
    )


def df_to_excel_bytes(
    df: pd.DataFrame,
    *,
    sheet_name: str = "Produtos",
) -> bytes:
    """
    Exporta o DataFrame exatamente como recebido, preservando ordem das colunas
    e nomes originais, sem adicionar campos extras.
    """

    if df is None:
        df = pd.DataFrame()

    df = df.copy()
    df = limpar_valores_vazios(df)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        if ws.max_row >= 1:
            for cell in ws[1]:
                header = str(cell.value or "").strip()
                if not header:
                    continue

                if _coluna_parece_codigo(header):
                    for body_cell in ws[cell.column_letter]:
                        body_cell.number_format = numbers.FORMAT_TEXT

    output.seek(0)
    return output.getvalue()
