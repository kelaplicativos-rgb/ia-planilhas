from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st

from bling_app_zero.core.audit import add_audit_event
from bling_app_zero.core.exporter import enforce_export_contract
from bling_app_zero.core.files import read_uploaded_file
from bling_app_zero.core.final_csv_exporter import exact_contract_columns

RESPONSIBLE_FILE = 'bling_app_zero/ui/exact_template_file_runtime.py'
PATCH_ATTR = '_exact_template_file_runtime_v1'
ORIGINAL_BYTES_ATTR = '_exact_template_original_to_bling_csv_bytes'
ORIGINAL_FILENAME_ATTR = '_exact_template_original_filename_for_operation'
MODEL_BYTES_KEY = 'destination_model_upload_bytes'
MODEL_NAME_KEY = 'destination_model_upload_name'

EXCEL_SUFFIXES = {'.xlsx', '.xlsm'}


class NamedBytesIO(BytesIO):
    def __init__(self, data: bytes, name: str):
        super().__init__(data)
        self.name = name


def _model_bytes_and_name() -> tuple[bytes, str]:
    data = st.session_state.get(MODEL_BYTES_KEY)
    name = str(st.session_state.get(MODEL_NAME_KEY) or '').strip()
    if isinstance(data, (bytes, bytearray)) and data and name:
        return bytes(data), name
    return b'', ''


def _model_suffix() -> str:
    _data, name = _model_bytes_and_name()
    return Path(name).suffix.lower()


def _model_columns_from_bytes(data: bytes, name: str) -> list[str]:
    try:
        df_model = read_uploaded_file(NamedBytesIO(data, name))
    except Exception:
        return []
    return exact_contract_columns(df_model.columns) if isinstance(df_model, pd.DataFrame) and len(df_model.columns) > 0 else []


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    if source_row <= 0 or target_row <= 0 or source_row == target_row:
        return
    try:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    except Exception:
        pass
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        try:
            if src.has_style:
                dst._style = copy(src._style)
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
        except Exception:
            pass


def _xlsx_from_template(data: bytes, name: str, df: pd.DataFrame, columns: list[str]) -> bytes:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data))
    ws = wb.active
    final_df = enforce_export_contract(df, columns).fillna('')
    max_col = len(columns)
    for col_idx, header in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx).value = header

    template_style_row = 2 if ws.max_row >= 2 else 1
    required_last_row = max(1 + len(final_df), ws.max_row)
    for row_idx in range(2, required_last_row + 1):
        if row_idx > ws.max_row:
            ws.append([''] * max_col)
        _copy_row_style(ws, template_style_row, row_idx, max_col)
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_idx, column=col_idx).value = ''

    for row_offset, values in enumerate(final_df[columns].astype(str).values.tolist(), start=2):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_offset, column=col_idx).value = '' if value == 'nan' else value

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def install_exact_template_file_runtime() -> bool:
    try:
        from bling_app_zero.ui import home_download
    except Exception:
        return False
    if getattr(home_download, PATCH_ATTR, False):
        return False

    original_bytes = getattr(home_download, ORIGINAL_BYTES_ATTR, None)
    if original_bytes is None:
        original_bytes = home_download.to_bling_csv_bytes
        setattr(home_download, ORIGINAL_BYTES_ATTR, original_bytes)

    original_filename = getattr(home_download, ORIGINAL_FILENAME_ATTR, None)
    if original_filename is None:
        original_filename = home_download.filename_for_operation
        setattr(home_download, ORIGINAL_FILENAME_ATTR, original_filename)

    def filename_for_operation(operation: str) -> str:
        _data, name = _model_bytes_and_name()
        if name and _model_suffix() in EXCEL_SUFFIXES:
            return name
        return original_filename(operation)

    def to_bling_csv_bytes(df: pd.DataFrame, operation: str = 'global', contract_columns=None, explicit_empty_columns=None) -> bytes:
        data, name = _model_bytes_and_name()
        suffix = Path(name).suffix.lower()
        if data and suffix in EXCEL_SUFFIXES:
            columns = exact_contract_columns(contract_columns) or _model_columns_from_bytes(data, name)
            if columns:
                add_audit_event('exact_template_xlsx_download_applied', area='DOWNLOAD', status='OK', details={'model_name': name, 'columns_count': len(columns), 'columns': columns, 'responsible_file': RESPONSIBLE_FILE})
                return _xlsx_from_template(data, name, df, columns)
        return original_bytes(df, operation=operation, contract_columns=contract_columns, explicit_empty_columns=explicit_empty_columns)

    home_download.filename_for_operation = filename_for_operation
    home_download.to_bling_csv_bytes = to_bling_csv_bytes
    setattr(home_download, PATCH_ATTR, True)
    add_audit_event('exact_template_file_runtime_installed', area='DOWNLOAD', status='OK', details={'responsible_file': RESPONSIBLE_FILE})
    return True


__all__ = ['install_exact_template_file_runtime']
