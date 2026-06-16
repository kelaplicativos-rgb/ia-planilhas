from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any

import csv
import hashlib
import zipfile

import pandas as pd
import streamlit as st

from bling_app_zero.core.audit import add_audit_event
from bling_app_zero.ui.home_shared import read_upload_fast
from bling_app_zero.universal.model_contract_detector import (
    MODEL_CONTRACT_CONFIDENCE_KEY,
    MODEL_CONTRACT_LABEL_KEY,
    MODEL_CONTRACT_REASON_KEY,
    MODEL_CONTRACT_TYPE_KEY,
    detect_model_contract,
)

MODEL_SPREADSHEET_TYPES = ['xlsx', 'xls', 'csv', 'xlsm', 'xlsb', 'zip']
EXCEL_HEADER_FALLBACK_TYPES = {'xlsx', 'xls', 'xlsm', 'xlsb'}
HEADER_FALLBACK_TYPES = EXCEL_HEADER_FALLBACK_TYPES | {'csv'}
MODEL_LABEL = 'Modelo para mapear'
MODEL_SIGNATURE_KEY = 'destination_model_upload_signature_current'
MODEL_BYTES_KEY = 'destination_model_upload_bytes'
MODEL_NAME_KEY = 'destination_model_upload_name'
MODEL_OBJECT_KEY = 'destination_model_upload_object'

MODEL_STATE_KEYS = (
    MODEL_SIGNATURE_KEY,
    MODEL_BYTES_KEY,
    MODEL_NAME_KEY,
    MODEL_OBJECT_KEY,
    'destination_model_contract_type',
    'destination_model_contract_label',
    'destination_model_contract_confidence',
    'destination_model_contract_reason',
    'model_contract_type',
    'model_contract_label',
    'model_contract_confidence',
    'model_contract_reason',
    'df_modelo_universal',
    'home_modelo_universal_df',
    'modelo_universal_df',
    'mapeiaai_final_contract_df',
    'cadastro_wizard_df_modelo',
    'home_modelo_cadastro_df',
    'df_modelo_cadastro',
    'modelo_cadastro_df',
    'home_modelo_estoque_df',
    'df_modelo_estoque',
    'modelo_estoque_df',
    'estoque_wizard_df_modelo',
    'home_modelo_atualizacao_preco_df',
    'df_modelo_atualizacao_preco',
    'modelo_atualizacao_preco_df',
    'home_model_upload_signature_v2',
    'home_model_save_signature_v3',
    'home_model_upload_autoforwarded_signature_v2',
)

OUTPUT_STATE_KEYS = (
    'df_final_universal',
    'df_final_cadastro',
    'df_final_cadastro_preview_rules_applied',
    'df_final_download_snapshot',
    'final_download_df_snapshot',
    'final_download_file_bytes',
    'final_download_file_name',
    'final_download_mime',
    'final_download_signature',
    'final_download_rules_signature',
    'final_download_operation',
    'final_download_widget_key',
)


@dataclass
class ModelUploadResult:
    cadastro_model_file: Any | None = None
    cadastro_model_df: pd.DataFrame | None = None
    estoque_model_file: Any | None = None
    estoque_model_df: pd.DataFrame | None = None
    model_file: Any | None = None
    model_df: pd.DataFrame | None = None
    attachments: list[Any] | None = None
    ignored_files: list[Any] | None = None
    contract_type: str = 'universal'
    contract_label: str = MODEL_LABEL
    contract_confidence: float = 0.0
    contract_reason: str = ''


def _file_name(file: Any | None) -> str:
    return str(getattr(file, 'name', 'arquivo')).strip() if file is not None else ''


def _file_ext(file: Any | None) -> str:
    name = _file_name(file).lower()
    return name.rsplit('.', 1)[-1] if '.' in name else ''


def _file_bytes(file: Any | None) -> bytes:
    if file is None:
        return b''
    try:
        return bytes(file.getvalue())
    except Exception:
        try:
            pos = file.tell()
            file.seek(0)
            data = file.read()
            file.seek(pos)
            return bytes(data or b'')
        except Exception:
            return b''


def _file_signature(file: Any | None) -> str:
    data = _file_bytes(file)
    if not data:
        return ''
    return f'{_file_name(file)}:{len(data)}:{hashlib.sha256(data).hexdigest()[:16]}'


def _has_previous_model_state() -> bool:
    return any(key in st.session_state for key in MODEL_STATE_KEYS)


def _clear_previous_model_state(reason: str) -> None:
    for key in MODEL_STATE_KEYS + OUTPUT_STATE_KEYS:
        st.session_state.pop(key, None)
    add_audit_event('mapping_model_previous_state_cleared', area='MODELO', status='OK', details={'reason': reason})


def _file_audit_info(file: Any | None) -> dict[str, Any] | None:
    if file is None:
        return None
    return {'name': _file_name(file), 'extension': _file_ext(file), 'size': getattr(file, 'size', None), 'type': getattr(file, 'type', None)}


def _df_audit_info(df: pd.DataFrame | None) -> dict[str, Any]:
    if not isinstance(df, pd.DataFrame):
        return {'valid': False}
    return {'valid': True, 'rows': int(len(df)), 'columns_count': int(len(df.columns)), 'columns': [str(column) for column in list(df.columns)[:120]]}


def _clean_header_value(value: Any) -> str:
    text = str(value or '').replace('\ufeff', '').replace('\xa0', ' ').strip()
    text = ' '.join(text.split())
    if not text or text.lower().startswith('unnamed:'):
        return ''
    return text


def _dedupe_columns(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    fixed: list[str] = []
    for index, column in enumerate(columns):
        base = _clean_header_value(column) or f'Coluna {index + 1}'
        count = seen.get(base, 0)
        seen[base] = count + 1
        fixed.append(base if count == 0 else f'{base} ({count + 1})')
    return fixed


def _valid_columns(columns: list[Any]) -> list[str]:
    return [_clean_header_value(value) for value in columns if _clean_header_value(value)]


def _header_df(columns: list[Any], *, source: str = '') -> pd.DataFrame | None:
    fixed = _dedupe_columns(_valid_columns(list(columns or [])))
    if not fixed:
        return None
    df = pd.DataFrame(columns=fixed)
    if source:
        df.attrs['model_header_source'] = source
    return df


def _detect_csv_separator(text: str) -> str:
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=',;\t|').delimiter
    except Exception:
        candidates = [',', ';', '\t', '|']
        return max(candidates, key=lambda sep: sample.count(sep))


def _decode_text_bytes(data: bytes) -> str:
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin1'):
        try:
            return data.decode(encoding)
        except Exception:
            continue
    return data.decode('utf-8', errors='ignore')


def _read_csv_header_from_bytes(data: bytes, file_name: str = 'modelo.csv') -> pd.DataFrame | None:
    text = _decode_text_bytes(data)
    if not text.strip():
        return None
    sep = _detect_csv_separator(text)
    try:
        df = pd.read_csv(StringIO(text), sep=sep, dtype=str, nrows=0)
        return _header_df(list(df.columns), source=file_name)
    except Exception:
        first_line = next((line for line in text.splitlines() if line.strip()), '')
        if not first_line:
            return None
        return _header_df([part.strip() for part in first_line.split(sep)], source=f'{file_name}:linha_1')


def _read_excel_header_with_pandas(data: bytes, file_name: str = 'modelo.xlsx') -> pd.DataFrame | None:
    try:
        frames = pd.read_excel(BytesIO(data), sheet_name=None, dtype=str, nrows=0)
    except Exception:
        return None
    best_df: pd.DataFrame | None = None
    best_sheet = ''
    for sheet_name, frame in frames.items():
        candidate = _header_df(list(frame.columns), source=f'{file_name}:{sheet_name}')
        if _valid_model(candidate) and (best_df is None or len(candidate.columns) > len(best_df.columns)):
            best_df = candidate
            best_sheet = str(sheet_name)
    if _valid_model(best_df):
        best_df.attrs['model_header_sheet'] = best_sheet
        return best_df
    return None


def _read_excel_header_with_openpyxl(data: bytes, file_name: str = 'modelo.xlsx') -> pd.DataFrame | None:
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    workbook = None
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        best_columns: list[str] = []
        best_sheet = ''
        best_row = 0
        for sheet in workbook.worksheets:
            max_rows_to_scan = min(int(sheet.max_row or 1), 30)
            max_cols_to_scan = int(sheet.max_column or 0)
            for row_index in range(1, max_rows_to_scan + 1):
                values = [sheet.cell(row=row_index, column=col_index).value for col_index in range(1, max_cols_to_scan + 1)]
                columns = _valid_columns(values)
                if len(columns) > len(best_columns):
                    best_columns = columns
                    best_sheet = sheet.title
                    best_row = row_index
        df = _header_df(best_columns, source=f'{file_name}:{best_sheet}:linha_{best_row}')
        if _valid_model(df):
            df.attrs['model_header_sheet'] = best_sheet
            df.attrs['model_header_row'] = best_row
        return df
    except Exception:
        return None
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass


def _read_header_from_bytes(data: bytes, file_name: str) -> pd.DataFrame | None:
    lower = str(file_name or '').lower()
    if lower.endswith('.csv'):
        return _read_csv_header_from_bytes(data, file_name)
    if lower.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
        pandas_df = _read_excel_header_with_pandas(data, file_name)
        if _valid_model(pandas_df):
            return pandas_df
        return _read_excel_header_with_openpyxl(data, file_name)
    return None


def _read_zip_header_fallback(file: Any) -> pd.DataFrame | None:
    try:
        with zipfile.ZipFile(BytesIO(_file_bytes(file))) as archive:
            best_df: pd.DataFrame | None = None
            best_inner = ''
            for info in archive.infolist():
                if info.is_dir():
                    continue
                inner_name = str(info.filename or '')
                lower = inner_name.lower()
                if not lower.endswith(('.csv', '.xlsx', '.xls', '.xlsm', '.xlsb')):
                    continue
                candidate = _read_header_from_bytes(archive.read(info), inner_name)
                if _valid_model(candidate) and (best_df is None or len(candidate.columns) > len(best_df.columns)):
                    best_df = candidate
                    best_inner = inner_name
            if _valid_model(best_df):
                best_df.attrs['model_header_zip_member'] = best_inner
                return best_df
    except Exception:
        return None
    return None


def _read_model_header_fallback(file: Any) -> pd.DataFrame | None:
    ext = _file_ext(file)
    try:
        if ext == 'zip':
            df = _read_zip_header_fallback(file)
        elif ext in HEADER_FALLBACK_TYPES:
            df = _read_header_from_bytes(_file_bytes(file), _file_name(file))
        else:
            df = None
        if _valid_model(df):
            add_audit_event('mapping_model_header_fallback_applied', area='MODELO', status='OK', details={'file': _file_audit_info(file), 'dataframe': _df_audit_info(df), 'source': getattr(df, 'attrs', {})})
            return df
        add_audit_event('mapping_model_header_fallback_empty', area='MODELO', status='AVISO', details={'file': _file_audit_info(file)})
        return None
    except Exception as exc:
        add_audit_event('mapping_model_header_fallback_failed', area='MODELO', status='ERRO', details={'file': _file_audit_info(file), 'error': str(exc)})
        return None


def _safe_read(file: Any) -> pd.DataFrame | None:
    fallback_df: pd.DataFrame | None = None
    try:
        df = read_upload_fast(file)
        if _valid_model(df):
            add_audit_event('mapping_model_file_read', area='MODELO', details={'file': _file_audit_info(file), 'dataframe': _df_audit_info(df), 'reader': 'read_upload_fast'})
            return df
        fallback_df = _read_model_header_fallback(file)
        if _valid_model(fallback_df):
            add_audit_event('mapping_model_file_read_recovered_by_header_fallback', area='MODELO', status='OK', details={'file': _file_audit_info(file), 'dataframe': _df_audit_info(fallback_df), 'reader': 'header_fallback_after_invalid_df'})
            return fallback_df
        add_audit_event('mapping_model_file_read', area='MODELO', details={'file': _file_audit_info(file), 'dataframe': _df_audit_info(df), 'reader': 'read_upload_fast_invalid'})
        return df
    except Exception as exc:
        fallback_df = _read_model_header_fallback(file)
        if _valid_model(fallback_df):
            add_audit_event('mapping_model_file_read_recovered_by_header_fallback', area='MODELO', status='OK', details={'file': _file_audit_info(file), 'dataframe': _df_audit_info(fallback_df), 'original_error': str(exc), 'reader': 'header_fallback_after_exception'})
            return fallback_df
        add_audit_event('mapping_model_file_read_failed', area='MODELO', status='ERRO', details={'file': _file_audit_info(file), 'error': str(exc)})
        return None


def _valid_model(df: pd.DataFrame | None) -> bool:
    return isinstance(df, pd.DataFrame) and len(df.columns) > 0


def _pick_destination_model(loaded: list[tuple[Any, pd.DataFrame | None]]) -> tuple[Any | None, pd.DataFrame | None]:
    for file, df in loaded:
        if _valid_model(df):
            return file, df.copy().fillna('')
    return None, None


def _columns_caption(df: pd.DataFrame, limit: int = 18) -> str:
    columns = [str(column) for column in list(df.columns)[:limit]]
    suffix = '...' if len(df.columns) > limit else ''
    return ', '.join(columns) + suffix


def _remember_contract_detection(detection) -> None:
    st.session_state[MODEL_CONTRACT_TYPE_KEY] = 'universal'
    st.session_state[MODEL_CONTRACT_LABEL_KEY] = MODEL_LABEL
    st.session_state[MODEL_CONTRACT_CONFIDENCE_KEY] = 1.0
    st.session_state[MODEL_CONTRACT_REASON_KEY] = 'Modelo para mapear enviado pelo usuário.'
    st.session_state['home_detected_operation'] = 'universal'


def _render_model_summary(file: Any | None, df: pd.DataFrame | None, detection=None) -> None:
    if not isinstance(df, pd.DataFrame):
        return
    st.success(f'{MODEL_LABEL} anexado.')
    st.caption(f'{_file_name(file)} · {len(df)} linha(s) · {len(df.columns)} coluna(s)')
    with st.expander('Ver colunas do modelo para mapear', expanded=False):
        st.caption(_columns_caption(df))


def _classified_result(destination_file: Any, destination_df: pd.DataFrame, detection, supported_files: list[Any], ignored_files: list[Any]) -> ModelUploadResult:
    df = destination_df.copy().fillna('')
    return ModelUploadResult(
        cadastro_model_file=destination_file,
        cadastro_model_df=df,
        estoque_model_file=destination_file,
        estoque_model_df=df,
        model_file=destination_file,
        model_df=df,
        attachments=supported_files,
        ignored_files=ignored_files,
        contract_type='universal',
        contract_label=MODEL_LABEL,
        contract_confidence=1.0,
        contract_reason='Modelo para mapear enviado pelo usuário.',
    )


def render_model_upload_box(title: str, operation: str, key: str, required_model: bool = False, caption: str | None = None) -> ModelUploadResult:
    files = st.file_uploader('Enviar modelo para mapear', type=None, accept_multiple_files=True, key=key, help='Envie a planilha modelo que será preenchida no final. Ela precisa ter cabeçalho com os nomes das colunas.', label_visibility='collapsed')

    if not files:
        if _has_previous_model_state():
            _clear_previous_model_state('modelo_removido_ou_uploader_vazio')
        return ModelUploadResult(attachments=[], ignored_files=[])

    selected_files = list(files)
    supported_files = [file for file in selected_files if _file_ext(file) in MODEL_SPREADSHEET_TYPES]
    ignored_files = [file for file in selected_files if _file_ext(file) not in MODEL_SPREADSHEET_TYPES]

    add_audit_event('mapping_model_upload_received', area='MODELO', details={'title': title, 'operation': operation, 'key': key, 'required_model': required_model, 'caption': caption, 'supported_count': len(supported_files), 'ignored_count': len(ignored_files), 'files': [_file_audit_info(file) for file in selected_files], 'model_policy': 'universal_mapping_model'})

    if not supported_files:
        _clear_previous_model_state('modelo_incompativel')
        st.warning('Nenhuma planilha compatível encontrada. Use XLSX, XLS, CSV, XLSM, XLSB ou ZIP com CSV/XLSX dentro.')
        return ModelUploadResult(attachments=[], ignored_files=ignored_files)

    current_signature = '|'.join(_file_signature(file) for file in supported_files)
    previous_signature = str(st.session_state.get(MODEL_SIGNATURE_KEY) or '')
    if current_signature and current_signature != previous_signature:
        _clear_previous_model_state('novo_modelo_anexado')
        st.session_state[MODEL_SIGNATURE_KEY] = current_signature

    with st.spinner('Lendo modelo para mapear...'):
        loaded = [(file, _safe_read(file)) for file in supported_files]
        destination_file, destination_df = _pick_destination_model(loaded)

    if not _valid_model(destination_df):
        _clear_previous_model_state('modelo_sem_colunas_validas')
        st.warning('Não encontrei colunas válidas nesse arquivo. Verifique se a planilha possui cabeçalho na primeira linha ou envie o modelo correto que será preenchido no final.')
        return ModelUploadResult(attachments=supported_files, ignored_files=ignored_files)

    st.session_state[MODEL_BYTES_KEY] = _file_bytes(destination_file)
    st.session_state[MODEL_NAME_KEY] = _file_name(destination_file)
    st.session_state[MODEL_OBJECT_KEY] = destination_file

    detection = detect_model_contract(destination_df)
    _remember_contract_detection(detection)
    add_audit_event('mapping_model_selected', area='MODELO', status='OK', details={'selected_model_file': _file_audit_info(destination_file), 'selected_df': _df_audit_info(destination_df), 'model_policy': 'universal_mapping_model', 'contract_type': 'universal', 'contract_label': MODEL_LABEL, 'signature': current_signature})
    _render_model_summary(destination_file, destination_df, detection)

    return _classified_result(destination_file, destination_df, detection, supported_files, ignored_files)


__all__ = ['MODEL_SPREADSHEET_TYPES', 'ModelUploadResult', 'render_model_upload_box']
